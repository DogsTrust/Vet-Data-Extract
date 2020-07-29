from datetime import datetime as dt
import os
import re
import pandas as pd
from openpyxl import load_workbook

#function to isolate ID numbers
def id(lst):
    output = []
    for id in lst:
        output.append(id[1:len(i)])
    return output

#Function to isolate age
def age_fn(lst):
    output = []
    for id in lst:
        output.append(id)
    return output

def age_in_months(age_in):
    ages_years = []
    for ages in age_in:
        yr = (re.search('\dyr', ages)).group()
        year_num = int((re.search('\d', yr)).group())*12
        mn = (re.search('\dm', ages)).group()
        month_num = (int((re.search('\d', mn)).group()))
        age_num = year_num + month_num
        ages_years.append(age_num)
    return ages_years

#Drive and file paths
Drive = 'M:\\'
input_path = os.path.join(Drive,'PDF Data Extraction','Vet Record Examples (1)','Text')
master_excel = os.path.join(Drive, 'PDF Data Extraction','Vet Record Examples (1)','Master_Record.xlsx')
filelist = os.listdir(input_path)


owner = []
vet = []
dog = []
age = []
for i in filelist:
    owner.append((re.search('O\d+', str(i))).group())
    vet.append((re.search('V\d+', str(i))).group())
    dog.append((re.search('D\d+', str(i))).group())
    age.append((re.search('\dyrs\s\dm', str(i))).group())

owner_id = id(owner)
vet_id = id(vet)
dog_id = id(dog)
ages = age_fn(age)
month_age = age_in_months(ages)
print(ages)

data = {'DogId':dog_id, 'OwnerId':owner_id, 'VetId':vet_id, 'Age (Text)':ages, 'Age (Months)':month_age}
df= pd.DataFrame(data)

#write to master doc
book = load_workbook(master_excel)
writer = pd.ExcelWriter(master_excel, engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}
for sheetname in writer.sheets:
    df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

writer.save()



